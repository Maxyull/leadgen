import csv

from leadgen.export import COLONNES, exporter, exporter_csv, exporter_xlsx

LIGNES = [
    {"email": "contact@cabinet-crea.fr", "nom": "CABINET CREA", "segment": "avocats",
     "type_email": "generique", "score": 90, "ville": "NANTES", "code_postal": "44000",
     "departement": "44", "tranche_effectif": "12", "naf": "69.10Z",
     "siren": "111111111", "site": "https://cabinet-crea.fr",
     "source_url": "https://cabinet-crea.fr/contact", "collecte_le": "2026-08-01T10:00:00"},
]


def test_csv_lisible_par_excel_fr(tmp_path):
    chemin = tmp_path / "leads.csv"
    assert exporter_csv(LIGNES, chemin) == 1
    brut = chemin.read_bytes()
    assert brut.startswith(b"\xef\xbb\xbf")          # BOM : accents corrects sous Excel
    texte = brut.decode("utf-8-sig")
    assert texte.splitlines()[0].count(";") == len(COLONNES) - 1
    lignes = list(csv.reader(texte.splitlines(), delimiter=";"))
    assert lignes[1][0] == "contact@cabinet-crea.fr"


def test_xlsx_avec_entete_figee(tmp_path):
    from openpyxl import load_workbook

    chemin = tmp_path / "leads.xlsx"
    assert exporter_xlsx(LIGNES, chemin) == 1
    ws = load_workbook(chemin).active
    assert ws["A1"].value == "Email"
    assert ws["A2"].value == "contact@cabinet-crea.fr"
    assert ws.freeze_panes == "A2"


def test_choix_du_format_par_extension(tmp_path):
    exporter(LIGNES, tmp_path / "a.csv")
    exporter(LIGNES, tmp_path / "b.xlsx")
    assert (tmp_path / "a.csv").exists()
    assert (tmp_path / "b.xlsx").exists()


def test_colonnes_personnalisees(tmp_path):
    chemin = tmp_path / "court.csv"
    exporter_csv(LIGNES, chemin, colonnes=["email", "nom"])
    assert chemin.read_text(encoding="utf-8-sig").splitlines()[0] == "Email;Structure"


def test_export_vide(tmp_path):
    chemin = tmp_path / "vide.csv"
    assert exporter_csv([], chemin) == 0
    assert chemin.exists()
