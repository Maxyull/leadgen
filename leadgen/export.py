"""Export du fichier de prospection (CSV ou XLSX).

Le CSV est encode en UTF-8 avec BOM : c'est le seul format qu'Excel FR ouvre
sans casser les accents, et la plupart des outils d'emailing l'acceptent.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Sequence

COLONNES = [
    "email", "nom", "segment", "type_email", "score", "ville", "code_postal",
    "departement", "tranche_effectif", "naf", "siren", "site", "source_url",
    "source", "titre", "accroche", "signaux", "angle", "collecte_le",
]

ENTETES_LISIBLES = {
    "email": "Email",
    "nom": "Structure",
    "segment": "Segment",
    "type_email": "Type d'adresse",
    "score": "Score",
    "ville": "Ville",
    "code_postal": "CP",
    "departement": "Dept",
    "tranche_effectif": "Effectif",
    "naf": "NAF",
    "siren": "SIREN",
    "site": "Site",
    "source_url": "Page source",
    "source": "Origine",
    "titre": "Titre du site",
    "accroche": "Ce qu'ils disent d'eux",
    "signaux": "Signaux",
    "angle": "Angle suggere",
    "collecte_le": "Collecte le",
}


def _lignes(rows: Iterable, colonnes: Sequence[str]) -> list[list]:
    sorties = []
    for r in rows:
        acces = r.keys() if hasattr(r, "keys") else r
        sorties.append([
            (r[c] if c in acces else "") if hasattr(r, "keys") else r.get(c, "")
            for c in colonnes
        ])
    return sorties


def exporter_csv(rows: Iterable, chemin: Path | str,
                 colonnes: Sequence[str] = COLONNES) -> int:
    chemin = Path(chemin)
    chemin.parent.mkdir(parents=True, exist_ok=True)
    lignes = _lignes(rows, colonnes)
    with chemin.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([ENTETES_LISIBLES.get(c, c) for c in colonnes])
        writer.writerows(lignes)
    return len(lignes)


def exporter_xlsx(rows: Iterable, chemin: Path | str,
                  colonnes: Sequence[str] = COLONNES) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    chemin = Path(chemin)
    chemin.parent.mkdir(parents=True, exist_ok=True)
    lignes = _lignes(rows, colonnes)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append([ENTETES_LISIBLES.get(c, c) for c in colonnes])
    for cellule in ws[1]:
        cellule.font = Font(bold=True, color="FFFFFF")
        cellule.fill = PatternFill("solid", fgColor="1F2937")
        cellule.alignment = Alignment(horizontal="center")
    for ligne in lignes:
        ws.append(ligne)
    ws.freeze_panes = "A2"
    largeurs = {"email": 34, "nom": 40, "site": 34, "source_url": 40, "segment": 20,
                "titre": 40, "accroche": 60, "signaux": 24, "angle": 50}
    for i, col in enumerate(colonnes, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largeurs.get(col, 14)
    if lignes:
        ws.auto_filter.ref = ws.dimensions
    wb.save(chemin)
    return len(lignes)


def exporter(rows: Iterable, chemin: Path | str,
             colonnes: Sequence[str] = COLONNES) -> int:
    rows = list(rows)
    if str(chemin).lower().endswith(".xlsx"):
        return exporter_xlsx(rows, chemin, colonnes)
    return exporter_csv(rows, chemin, colonnes)
